#!/usr/sbin/python3

from openpyxl import Workbook, load_workbook
from xlcalculator import ModelCompiler, Model, Evaluator
from base64 import b64encode
import uuid
import os
import random
from secret import flag, content

class Wishlist(object):
    def __init__(self, wb, ws):
        self.wb = wb
        self.ws = ws
        self.content = {}
        self.ws.cell(row=2, column=5, value=100)
        self.insert_item(3, 'flag', 1e100)
        self.content = {'flag':flag}
        self.ws.cell(row=4, column=5, value=f'=E2-SUM(E{3}:E{3})')
        self.lastrow = ws.max_row
        self.tmpfile = tmpfile = f'/tmp/{uuid.uuid4()}.xlsx'

    def insert_item(self, row, itemname, price):
        self.ws.cell(row=row, column=1, value=itemname)
        self.ws.cell(row=row, column=3, value=price)
        self.ws.cell(row=row, column=4, value=0)
        self.ws.cell(row=row, column=5, value=f'=C{row}*D{row}')
        self.ws.cell(row=row, column=6, value=0)
        self.content[itemname] = random.choice(content)

    def add_item(self, itemname, price):
        if price<0:
            print('Who in the right mind sells stuff at negative price?')
            return
        for i in range(3, self.lastrow):
            if self.ws.cell(row=i, column=1).value==itemname:
                print(f'{itemname} already exists')
                return
        self.insert_item(self.lastrow, itemname, price)
        self.lastrow+=1
        self.ws.cell(row=self.lastrow, column=5, value=f'=E2-SUM(E{3}:E{self.lastrow-1})')
        print(f'{itemname} added')
        return

    def add_bundle(self,bundlename,items):
        for i in range(3, self.lastrow):
            if self.ws.cell(row=i, column=1).value==bundlename:
                print(f'{bundlename} already exists')
                return
        EQ = '=SUM('
        for itemname in items:
            FOUND = False
            for i in range(3, self.lastrow):
                if self.ws.cell(row=i, column=1).value==itemname:
                    EQ+=f'C{i},'
                    FOUND = True
                    break
            if FOUND is False:
                print(f"Can't add non-existing {itemname} to bundle")
                return
        if EQ[-1]==',':
            EQ = EQ[:-1]
        EQ+=')'
        self.insert_item(self.lastrow, bundlename, EQ)
        self.lastrow+=1
        self.ws.cell(row=self.lastrow, column=5, value=f'=E2-SUM(E{3}:E{self.lastrow-1})')
        print(f'{bundlename} added')
        return

    def add_to_cart(self, itemname, itemcnt):
        if itemcnt<0:
            print('Trying to sell stuff instead?\nToo bad we caught you redhanded.')
            return
        for i in range(3, self.lastrow):
            if self.ws.cell(row=i, column=1).value==itemname:
                self.ws.cell(row=i, column=4, value=itemcnt)
                print(f'{itemname} x {itemcnt} added to Cart')
                return
        print(f'{itemname} not found')
        return

    def buy(self):
        self.save()
        compiler = ModelCompiler()
        new_model = compiler.read_and_parse_archive(self.tmpfile)
        self.remove()
        evaluator = Evaluator(new_model)
        balance = evaluator.evaluate(f'mywishlist!E{self.lastrow}')
        if balance<0:
            print('Poor you')
        else:
            self.ws.cell(row=2, column=5, value=float(balance))
            for i in range(3,self.lastrow):
                owned = self.ws.cell(row=i, column=6).value
                bought = self.ws.cell(row=i, column=4).value
                owned+=bought
                if owned!=0:
                    self.ws.cell(row=i, column=2, value=self.content[self.ws.cell(row=i, column=1).value])
                self.ws.cell(row=i, column=6, value=owned)
                self.ws.cell(row=i, column=4, value=0)
            print('Purchased')
        return

    def export(self):
        self.save()
        res = open(self.tmpfile,'rb').read()
        self.remove()
        print(f'BASE64 : {b64encode(res).decode()}')

    def save(self):
        self.wb.save(self.tmpfile)

    def remove(self):
        os.remove(self.tmpfile)

    @classmethod
    def load(cls, wb):
        return cls(wb, wb.active)

def myinput(prompt):
    '''
    python input prompts to stderr by default, and there is no option to change this afaik
    this wrapper is just normal input with stdout prompt
    '''
    print(prompt,end='')
    return input()

def menu():
    print('')
    print('========= M30W Shop =========')
    print('  1. Add item to wishlist')
    print('  2. Add bundle to wishlist')
    print('  3. Add item(s) to cart')
    print('  4. Purchase')
    print('  5. Export bought goods')
    print('  6. Leave')
    print('=============================')

if __name__=='__main__':
    wishlist = Wishlist.load(load_workbook(filename='/home/ExcellentM30WStore/Wishlist.xlsx'))
    print('Welcome to M30W shop')
    print('Fill in tour wishlist to buy cat related goods here')
    print('Or if you are wealthy enough, we also have a flag for you')
    while True:
        menu()
        option = int(myinput('Choice > ').strip())
        if option==1:
            itemname = myinput('Item name : ').strip()
            price = int(myinput('Price : ').strip())
            wishlist.add_item(itemname, price)
        elif option==2:
            bundlename = myinput('Bundle name : ').strip()
            itemlist = []
            cnt = 1
            print('Provide a list of items to put in bundle')
            while True:
                itemname = myinput(f'  Item {cnt} : ').strip()
                if itemname=='':
                    break
                itemlist.append(itemname)
                cnt+=1
            wishlist.add_bundle(bundlename, itemlist)
        elif option==3:
            itemname = myinput('Item name : ').strip()
            quantity = int(myinput('Quantity : ').strip())
            wishlist.add_to_cart(itemname, quantity)
        elif option==4:
            wishlist.buy()
        elif option==5:
            wishlist.export()
        elif option==6:
            print('Thanks for your patronage')
            print('We will be looking forward to serve you again soon')
            break
        else:
            print('Invalid option, what do you think you are doing?')
